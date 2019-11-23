from django.contrib.auth.forms import UserChangeForm, PasswordChangeForm
from django.shortcuts import render, reverse
from django.template import RequestContext
from django.shortcuts import render_to_response, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from propiedades2.models import Acquisition, DocumentEx, DocumentCip, DocumentCn, DocumentBlue, DocumentBuildP, \
    DocumentMR, DocumentTypeC, DocumentOther, DocumentWR, DocumentDC, DocumentPH, DocumentDB, DocumentAc, DocumentEs, \
    Rent, Location, Post, ArchitectureRecordAcq, InternalAccountantsAcq,NotaryAcquisition,SiiRecord, Staff, Region, Property, Change_property, \
    Comment, Stats, District, Notification, Report
from propiedades2.forms import DocCipForm, DocCnForm, DocBlueForm, DocBuildPForm, DocMRForm, DocTypeCForm, \
    DocOtherForm, DocWRForm, DocDCForm, DocPHForm, DocDBForm, DocAcForm, DocEsForm, LocationForm, AcquisitionForm, \
    ArquitectureForm, InternalForm, NotaryForm, SII_recordForm, RentForm, DocExForm, StaffForm, UserForm, EditStaffForm, \
    EditUserForm, PasswordChangeForm
from django.contrib.auth.models import User
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponseRedirect
from django.db.models import Q
from django.http import HttpResponse
from django.core.mail import send_mail
import random
import string
import reversion
import datetime
from django.forms.models import model_to_dict
from reversion.models import Version
from django.views.decorators.csrf import csrf_protect
import re
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.drawing.image import Image
from django.utils import timezone
from openpyxl.styles import Alignment,Border, Font, PatternFill, Side, Color, NamedStyle


# Create your views here.

def randomString(stringLength=4):
    """Generate a random string of fixed length """
    letters = string.ascii_lowercase
    global CODE
    CODE = ''.join(random.choice(letters) for i in range(stringLength))

def index(request):
    data = {}
    data['staff'] = Staff.objects.get(username_staff = request.user)
    return render(request, 'index.html', data)


@login_required(login_url='login')
def list_acquisition(request):
    data = {}
    object_list = Acquisition.objects.all().order_by('-id')
    paginator = Paginator(object_list, 10)
    page = request.GET.get('page')
    data['staff'] = Staff.objects.get(username_staff=request.user)
    try:
        data['object_list'] = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        data['object_list'] = paginator.page(1)
    except EmptyPage:
        data['object_list'] = paginator.page(paginator.num_pages)
    template_name = 'list_acquisition.html'
    return render (request, template_name, data)

@login_required(login_url='login')
def view_acquisition(request, cli_id):
    data = {}
    data['staff'] = Staff.objects.get(username_staff=request.user)
    template_name = 'detail_acquisition.html'
    data['acquisition'] = Acquisition.objects.get(pk=cli_id)
    data['post_list'] = Post.objects.filter(acquisition__pk = cli_id).order_by('-publish_date')
    data['comment_post'] = Comment.objects.filter(post__acquisition__pk = cli_id)
    if Change_property.objects.filter(id_property = cli_id, type= 'Acq').exists():
        print(Change_property.objects.all().filter(id_property = cli_id, type = 'Acq'))
        data['change'] = Change_property.objects.all().filter(id_property = cli_id, type = 'Acq')
        return render(request, template_name, data)
    else:
        return render(request, template_name, data)


@login_required(login_url='login')
def view_rent(request, rent_id):
    data = {}
    data['staff'] = Staff.objects.get(username_staff=request.user)
    template_name = 'detail_rent.html'
    data['rent'] = Rent.objects.get(pk=rent_id)
    data['post_list'] = Post.objects.filter(rent__pk=rent_id).order_by('-publish_date')
    data['comment_post'] = Comment.objects.filter(post__rent__pk=rent_id)
    if Change_property.objects.filter(id_property = rent_id, type= 'rent').exists():
        data['change'] = Change_property.objects.all().filter(id_property = rent_id, type = 'Rent')
        return render(request, template_name, data)
    else:
        return render(request, template_name, data)


@login_required(login_url='login')
def list_rent(request):
    data = {}
    object_list = Rent.objects.all().order_by('-id')
    paginator = Paginator(object_list, 10)
    page = request.GET.get('page')
    data['staff'] = Staff.objects.get(username_staff=request.user)
    print(request.user)
    print(data['staff'].first_name)
    try:
        data['object_list'] = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        data['object_list'] = paginator.page(1)
    except EmptyPage:
        data['object_list'] = paginator.page(paginator.num_pages)
    template_name = 'list_rent.html'
    return render(request, template_name, data)


@login_required(login_url='login')
def Add_acquisition(request):
    data = {}
    staff = Staff.objects.get(username_staff=request.user)
    if staff.type_user == 'DIG' or staff.type_user == 'ADM':
        if request.method == "POST":
            # ArquitectureForm,InternalForm,NotaryForm,SII_recordForm
            data['formAcquisition'] = AcquisitionForm(request.POST, request.FILES)
            data['formLocation'] = LocationForm(request.POST, request.FILES)
            # Arquitecture
            data['formArquitecture'] = ArquitectureForm(request.POST, request.FILES)
            data['formEx'] = DocExForm(request.POST, request.FILES)
            data['formCip'] = DocCipForm(request.POST, request.FILES)
            data['formCn'] = DocCnForm(request.POST, request.FILES)
            data['formBlue'] = DocBlueForm(request.POST, request.FILES)
            data['formBuildP'] = DocBuildPForm(request.POST, request.FILES)
            data['formMR'] = DocMRForm(request.POST, request.FILES)
            # Internal
            data['formInternal'] = InternalForm(request.POST, request.FILES)
            data['formTypeC'] = DocTypeCForm(request.POST, request.FILES)
            data['formOther'] = DocOtherForm(request.POST, request.FILES)
            # Notary
            data['formNotary'] = NotaryForm(request.POST, request.FILES)
            data['formWR'] = DocWRForm(request.POST, request.FILES)
            data['formDC'] = DocDCForm(request.POST, request.FILES)
            data['formPH'] = DocPHForm(request.POST, request.FILES)
            data['formEs'] = DocEsForm(request.POST, request.FILES)
            # Sii
            data['formSII'] = SII_recordForm(request.POST, request.FILES)
            data['formAc'] = DocAcForm(request.POST, request.FILES)
            data['formDB'] = DocDBForm(request.POST, request.FILES)
            if request.POST.get('Tipo') is not None:
                if request.POST['Tipo'] == 'Propiedad':
                    if data['formAcquisition'].is_valid():
                        data['formAcquisition'] = data['formAcquisition'].save(commit=False)
                        if data['formLocation'].is_valid():
                            print('paso formlocation')
                            data['formLocation'] = data['formLocation'].save()
                            aux = Location.objects.get(pk=data['formLocation'].pk)
                            data['formAcquisition'].location = aux
                            data['formAcquisition'].save()
                            create_notification(request.user, "AA", data['formAcquisition'].pk, 'AC')
                            return JsonResponse({'id': data['formAcquisition'].pk})
            if request.POST.get('Arquitecture') is not None:
                if request.POST['Arquitecture'] == 'arquitecture':
                    print(request.POST)
                    if request.POST.get('id_propiedad') is not None:
                        prop = Acquisition.objects.get(pk=request.POST.get('id_propiedad'))
                        id = request.POST.get('id_propiedad')
                        if data['formArquitecture'].is_valid():
                            data['formArquitecture'] = data['formArquitecture'].save(commit=False)
                        if data['formEx'].is_valid():
                            Ex = DocumentEx()
                            if ('ArDocEx' not in request.POST):
                                Ex.archive = request.FILES['ArDocEx']
                            Ex.comment = request.POST['CDocEx']
                            Ex.type = 'EM'
                            Ex.save()
                            data['formArquitecture'].expropriation_mun = Ex
                        if data['formCip'].is_valid():
                            Cip = DocumentCip()
                            if ('ArDocCip' not in request.POST):
                                Cip.archive = request.FILES['ArDocCip']
                            Cip.comment = request.POST['CDocCip']
                            Cip.type = 'CP'
                            Cip.save()
                            data['formArquitecture'].cip = Cip
                        if data['formCn'].is_valid():
                            cn = DocumentCn()
                            if ('ArDocCn' not in request.POST):
                                cn.archive = request.FILES['ArDocCn']
                            cn.comment = request.POST['CDocCn']
                            cn.type = 'NC'
                            cn.save()
                            data['formArquitecture'].certified_number = cn
                        if data['formBlue'].is_valid():
                            blue = DocumentBlue()
                            if ('ArDocBlue' not in request.POST):
                                blue.archive = request.FILES['ArDocBlue']
                            blue.comment = request.POST['CDocBlue']
                            blue.type = 'PL'
                            blue.save()
                            data['formArquitecture'].blueprints = blue
                        if data['formBuildP'].is_valid():
                            build = DocumentBuildP()
                            if ('ArDocBuild' not in request.POST):
                                build.archive = request.FILES['ArDocBuild']
                            build.comment = request.POST['CDocBuild']
                            build.type = 'PE'
                            build.save()
                            data['formArquitecture'].building_permit = build
                        if data['formMR'].is_valid():
                            MR = DocumentMR()
                            if ('ArDocMR' not in request.POST):
                                MR.archive = request.FILES['ArDocMR']
                            MR.type = 'RM'
                            MR.save()
                            data['formArquitecture'].municipal_reception = MR
                        data['formArquitecture'].save()
                        prop.arquitecture = data['formArquitecture']
                        prop.save()

                    return JsonResponse({})

            elif request.POST.get('Internal') is not None:
                if request.POST['Internal'] == 'internal':
                    if request.POST.get('id_propiedad') is not None:
                        # extraer pk de la acquisition
                        prop = Acquisition.objects.get(pk=request.POST['id_propiedad'])
                    if data['formInternal'].is_valid():
                        data['formInternal'] = data['formInternal'].save(commit=False)
                        if data['formTypeC'].is_valid():
                            TypeC = DocumentTypeC()
                            if ('ArDocTypeC' not in request.POST):
                                TypeC.archive = request.FILES['ArDocTypeC']
                            TypeC.comment = request.POST['CDocTypeC']
                            TypeC.type = 'TC'
                            TypeC.save()
                            data['formInternal'].contract_type = TypeC
                        if data['formOther'].is_valid():
                            Other = DocumentOther()
                            if ('ArDocOther' not in request.POST):
                                Other.archive = request.FILES['ArDocOther']
                            Other.comment = request.POST['CDocOther']
                            Other.type = 'OT'
                            Other.save()
                            data['formInternal'].others = Other
                            # guardar en acquisition el intenal
                        data['formInternal'].save()
                        prop.internal = data['formInternal']
                        prop.save()
                        return JsonResponse({})

            elif request.POST.get('Notary') is not None:
                if request.POST['Notary'] == 'notary':
                    if request.POST.get('id_propiedad') is not None:
                        # extraer pk de la acquisition
                        prop = Acquisition.objects.get(pk=request.POST['id_propiedad'])

                    if data['formNotary'].is_valid():
                        data['formNotary'] = data['formNotary'].save(commit=False)
                        if data['formWR'].is_valid():
                            WR = DocumentWR()
                            if ('ArDocWR' not in request.POST):
                                WR.archive = request.FILES['ArDocWR']
                            WR.comment = request.POST['CDocWR']
                            WR.type = 'ES'
                            WR.save()
                            data['formNotary'].writing = WR
                        if data['formDC'].is_valid():
                            DC = DocumentDC()
                            if ('ArDocDC' not in request.POST):
                                DC.archive = request.FILES['ArDocDC']
                            DC.comment = request.POST['CDocDC']
                            DC.type = 'DC'
                            DC.save()
                            data['formNotary'].domain_certificate = DC
                        if data['formPH'].is_valid():
                            PH = DocumentPH()
                            if ('ArDocPH' not in request.POST):
                                PH.archive = request.FILES['ArDocPH']
                            PH.comment = request.POST['CDocPH']
                            PH.type = 'PR'
                            PH.save()
                            data['formNotary'].prohibitions = PH
                        if data['formDB'].is_valid():
                            Es = DocumentEs()
                            if ('ArDocDB' not in request.POST):
                                Es.archive = request.FILES['ArDocDB']
                            Es.comment = request.POST['CDocDB']
                            Es.type = 'SE'
                            Es.save()
                            data['formNotary'].expropriation_serviu = Es
                        data['formNotary'].save()
                        prop.notary = data['formNotary']
                        prop.save()
                    return JsonResponse({})

            elif request.POST.get('SII') is not None:
                if request.POST['SII'] == 'Sii':
                    if request.POST.get('id_propiedad') is not None:
                        # extraer pk de la acquisition
                        prop = Acquisition.objects.get(pk=request.POST['id_propiedad'])
                    if data['formSII'].is_valid():
                        data['formSII'] = data['formSII'].save(commit=False)
                        if data['formAc'].is_valid():
                            Ac = DocumentAc()
                            if ('ArDocAc' not in request.POST):
                                Ac.archive = request.FILES['ArDocAc']
                            Ac.comment = request.POST['CDocAc']
                            Ac.type = 'CA'
                            Ac.save()
                            data['formSII'].appraisal_certificate = Ac
                        if data['formEs'].is_valid():
                            DB = DocumentDB()
                            if ('ArDocEs' not in request.POST):
                                DB.archive = request.FILES['ArDocEs']
                            DB.comment = request.POST['CDocEs']
                            DB.type = 'CD'
                            DB.save()
                            data['formSII'].debt_certificate = DB
                        data['formSII'].save()
                        prop.SII = data['formSII']
                        prop.save()
                        #valores[0] = total
                        #valores[1] = complete
                        valores = walk_object(prop.pk)
                        prop.stats_acquisition = create_stats(valores[0], valores[1])
                        prop.save()
                    return JsonResponse({})
        else:
            data = {
                'formLocation': LocationForm, 'formAcquisition': AcquisitionForm, 'formArquitecture': ArquitectureForm,
                'formEx': DocExForm, 'formCip': DocCipForm, 'formCn': DocCnForm,
                'formBlue': DocBlueForm,
                'formBuildP': DocBuildPForm, 'formMR': DocMRForm, 'formInternal': InternalForm,
                'formTypeC': DocTypeCForm,
                'formOther': DocOtherForm, 'formNotary': NotaryForm, 'formWR': DocWRForm,
                'formDC': DocDCForm,
                'formPH': DocPHForm, 'formEs': DocEsForm, 'formSII': SII_recordForm,
                'formAc': DocAcForm, 'formDB': DocDBForm, 'staff': staff,
            }

        return render(request, 'add_acquisition.html', data)
    else:
        data['alerta'] = True
        a=list_acquisition(request)
        print(a)
        return HttpResponseRedirect(reverse('list_acquisition'))

@login_required(login_url='login')
def Add_rent(request):
    data = {}
    data['staff'] = Staff.objects.get(username_staff=request.user)
    if data['staff'].type_user == 'DIG' or data['staff'].type_user == 'ADM':
        if request.method == "POST":
            data['formLocation'] = LocationForm(request.POST)
            data['formRent'] = RentForm(request.POST, request.FILES)
            data['formDocument'] = DocTypeCForm(request.POST, request.FILES)
            if data['formRent'].is_valid():
                rent = data['formRent'].save(commit=False)
                if data['formLocation'].is_valid():
                    location = data['formLocation'].save()
                    rent.location = location
                    if data['formDocument'].is_valid():
                        document = data['formDocument'].save()
                        rent.contract_type = document
                        total = 0
                        contestado = 0
                        if request.POST['plot'] == '':
                            print('vacio')
                            total -= 1
                        if request.POST['lot_number'] == '':
                            print('vacio2')
                            total -= 1
                        for val in request.POST:
                            if request.POST[val] != '':
                                contestado += 1
                                total += 1
                            else:
                                total += 1
                        
                            #y es donde se modifica el stats correspondiente a la propiedad
                    rent.stats_rent = create_stats(total-1, contestado-1)
                    rent.save()
                    create_notification(request.user, "AR", rent.pk, "RT")
                return redirect('list_rent')

        else:
            data['formDocument'] = DocTypeCForm()
            data['formLocation'] = LocationForm()
            data['formRent'] = RentForm()
    else:
        data['alerta'] = True
        return HttpResponseRedirect(reverse('list_rent'))

    return render(request, 'add_rent.html', data)

@login_required(login_url='login')
def Edit_acquisition(request, acq_id):
    data = {}
    propiedad = get_object_or_404(Acquisition, pk=acq_id)
    data['pk'] = acq_id
    data['staff'] = Staff.objects.get(username_staff=request.user)
    if data['staff'].type_user == 'DIG' or data['staff'].type_user == 'ADM':
        if request.method == "POST":
            # ArquitectureForm,InternalForm,NotaryForm,SII_recordForm
            print('----------------------------------')
            print(propiedad.arquitecture.expropriation_mun.pk)
            print('----------------------------------')
            data['formAcquisition'] = AcquisitionForm(request.POST, request.FILES, instance=Acquisition.objects.get(pk=acq_id))
            data['formLocation'] = LocationForm(request.POST, request.FILES, instance = propiedad.location)
            # Arquitecture
            data['formArquitecture'] = ArquitectureForm(request.POST, request.FILES, instance = propiedad.arquitecture)
            data['formEx'] = DocExForm(request.POST, request.FILES)
            data['formCip'] = DocCipForm(request.POST, request.FILES)
            data['formCn'] = DocCnForm(request.POST, request.FILES)
            data['formBlue'] = DocBlueForm(request.POST, request.FILES)
            data['formBuildP'] = DocBuildPForm(request.POST, request.FILES)
            data['formMR'] = DocMRForm(request.POST, request.FILES)
            # Internal
            data['formInternal'] = InternalForm(request.POST, request.FILES, instance = propiedad.internal)
            data['formTypeC'] = DocTypeCForm(request.POST, request.FILES)
            data['formOther'] = DocOtherForm(request.POST, request.FILES)
            # Notary
            data['formNotary'] = NotaryForm(request.POST, request.FILES, instance = propiedad.notary)
            data['formWR'] = DocWRForm(request.POST, request.FILES)
            data['formDC'] = DocDCForm(request.POST, request.FILES)
            data['formPH'] = DocPHForm(request.POST, request.FILES)
            data['formEs'] = DocEsForm(request.POST, request.FILES)
            # Sii
            data['formSII'] = SII_recordForm(request.POST, request.FILES, instance = propiedad.SII)
            data['formAc'] = DocAcForm(request.POST, request.FILES)
            data['formDB'] = DocDBForm(request.POST, request.FILES)
            data['id'] = acq_id


            if request.POST.get('Tipo') is not None:
                if request.POST['Tipo'] == 'Propiedad':
                    if data['formAcquisition'].is_valid():
                        data['formAcquisition'] = data['formAcquisition'].save(commit=False)
                        if data['formLocation'].is_valid():
                            data['formLocation'] = data['formLocation'].save()
                            data['formAcquisition'].location = data['formLocation']
                            data['formAcquisition'].save()
                            create_notification(request.user, "EA", acq_id, 'AC')
                            return JsonResponse({'id': data['formAcquisition'].pk})

            elif request.POST.get('Arquitecture') is not None:
                if request.POST['Arquitecture'] == 'arquitecture':
                    if request.POST.get('id_propiedad') is not None:
                        # extraer pk de la acquisition
                        prop = Acquisition.objects.get(pk=acq_id)
                    if data['formArquitecture'].is_valid():
                        data['formArquitecture'] = data['formArquitecture'].save()
                        if data['formEx'].is_valid():
                            Ex = DocumentEx.objects.get(pk=propiedad.arquitecture.expropriation_mun.pk)
                            if request.FILES.get('ArDocEx') != None:
                                Ex.archive = request.FILES.get('ArDocEx')
                            Ex.comment = request.POST['CDocEx']
                            Ex.type = 'EM'
                            Ex.save()
                            print(Ex)
                            data['formArquitecture'].expropriation_mun = Ex
                        if data['formCip'].is_valid():
                            print('paso CIP')
                            Cip = DocumentCip.objects.get(pk=propiedad.arquitecture.cip.pk)
                            if request.FILES.get('ArDocCip') != None:
                                Cip.archive = request.FILES.get('ArDocCip')
                            Cip.comment = request.POST['CDocCip']
                            Cip.type = 'CP'
                            Cip.save()
                            data['formArquitecture'].cip = Cip
                        if data['formCn'].is_valid():
                            print('paso CN')
                            cn = DocumentCn.objects.get(pk=propiedad.arquitecture.certified_number.pk)
                            if request.FILES.get('ArDocCn') != None:
                                cn.archive = request.FILES.get('ArDocCn')
                            cn.comment = request.POST['CDocCn']
                            cn.type = 'NC'
                            cn.save()
                            data['formArquitecture'].certified_number = cn
                        if data['formBlue'].is_valid():
                            print('paso BLUE')
                            blue = DocumentBlue.objects.get(pk=propiedad.arquitecture.blueprints.pk)
                            if request.FILES.get('ArDocBlue') != None:
                                blue.archive = request.FILES.get('ArDocBlue')
                            blue.comment = request.POST['CDocBlue']
                            blue.type = 'PL'
                            blue.save()
                            data['formArquitecture'].blueprints = blue
                        if data['formBuildP'].is_valid():
                            print('paso BP')
                            build = DocumentBuildP.objects.get(pk=propiedad.arquitecture.building_permit.pk)
                            if request.FILES.get('ArDocBuild') != None:
                                build.archive = request.FILES.get('ArDocBuild')
                            build.comment = request.POST['CDocBuild']
                            build.type = 'PE'
                            build.save()
                            data['formArquitecture'].building_permit = build
                        if data['formMR'].is_valid():
                            print('paso MR')
                            MR = DocumentMR.objects.get(pk = propiedad.arquitecture.municipal_reception.pk)
                            if request.FILES.get('ArDocMR') != None:
                                MR.archive = request.FILES.get('ArDocMR')
                            MR.type = 'RM'
                            MR.save()
                            data['formArquitecture'].municipal_reception = MR
                        data['formArquitecture'].save()
                        prop.arquitecture = data['formArquitecture']
                        prop.save()
                        valores = walk_object(acq_id)
                        prop.stats_acquisition.total = valores[0]
                        prop.stats_acquisition.complete = valores[1]
                        prop.save()
                    return JsonResponse({})

            elif request.POST.get('Internal') is not None:
                if request.POST['Internal'] == 'internal':
                    if request.POST.get('id_propiedad') is not None:
                        # extraer pk de la acquisition
                        prop = Acquisition.objects.get(pk=acq_id)
                    if data['formInternal'].is_valid():
                        data['formInternal'] = data['formInternal'].save(commit=False)
                        if data['formTypeC'].is_valid():
                            TypeC = DocumentTypeC.objects.get(pk = propiedad.internal.contract_type.pk)
                            if request.FILES.get('ArDocTypeC') != None:
                                TypeC.archive = request.FILES.get('ArDocTypeC')
                            TypeC.comment = request.POST['CDocTypeC']
                            TypeC.type = 'TC'
                            TypeC.save()
                            data['formInternal'].contract_type = TypeC
                        if data['formOther'].is_valid():
                            Other = DocumentOther.objects.get(pk = propiedad.internal.others.pk)
                            if request.FILES.get('ArDocOther') != None:
                                Other.archive = request.FILES.get('ArDocOther')
                            Other.comment = request.POST['CDocOther']
                            Other.type = 'OT'
                            Other.save()
                            data['formInternal'].others = Other
                            # guardar en acquisition el intenal
                        data['formInternal'].save()
                        prop.internal = data['formInternal']
                        prop.save()
                        valores = walk_object(acq_id)
                        prop.stats_acquisition.total = valores[0]
                        prop.stats_acquisition.complete = valores[1]
                        prop.save()
                        return JsonResponse({})

            elif request.POST.get('Notary') is not None:
                if request.POST['Notary'] == 'notary':
                    if request.POST.get('id_propiedad') is not None:
                        # extraer pk de la acquisition
                        prop = Acquisition.objects.get(pk=acq_id)

                    if data['formNotary'].is_valid():
                        data['formNotary'] = data['formNotary'].save(commit=False)
                        if data['formWR'].is_valid():
                            WR = DocumentWR.objects.get(pk = propiedad.notary.writing.pk)
                            if request.FILES.get('ArDocWR') != None:
                                WR.archive = request.FILES.get('ArDocWR')
                            WR.comment = request.POST['CDocWR']
                            WR.type = 'ES'
                            WR.save()
                            data['formNotary'].writing = WR
                        if data['formDC'].is_valid():
                            DC = DocumentDC.objects.get(pk = propiedad.notary.domain_certificate.pk)
                            if request.FILES.get('ArDocDC') != None:
                                DC.archive = request.FILES.get('ArDocDC')
                            DC.comment = request.POST['CDocDC']
                            DC.type = 'DC'
                            DC.save()
                            data['formNotary'].domain_certificate = DC
                        if data['formPH'].is_valid():
                            PH = DocumentPH.objects.get(pk = propiedad.notary.prohibitions.pk)
                            if request.FILES.get('ArDocPH') != None:
                                PH.archive = request.FILES.get('ArDocPH')
                            PH.comment = request.POST['CDocPH']
                            PH.type = 'PR'
                            PH.save()
                            data['formNotary'].prohibitions = PH
                        if data['formDB'].is_valid():
                            Es = DocumentEs.objects.get(pk = propiedad.notary.expropriation_serviu.pk)
                            if request.FILES.get('ArDocDB') != None:
                                Es.archive = request.FILES.get('ArDocDB')
                            Es.comment = request.POST['CDocDB']
                            Es.type = 'SE'
                            Es.save()
                            data['formNotary'].expropriation_serviu = Es
                        data['formNotary'].save()
                        prop.notary = data['formNotary']
                        prop.save()
                        valores = walk_object(acq_id)
                        prop.stats_acquisition.total = valores[0]
                        prop.stats_acquisition.complete = valores[1]
                        prop.save()
                    return JsonResponse({})

            elif request.POST.get('SII') is not None:
                if request.POST['SII'] == 'Sii':
                    if request.POST.get('id_propiedad') is not None:
                        # extraer pk de la acquisition
                        prop = Acquisition.objects.get(pk=acq_id)
                    if data['formSII'].is_valid():
                        data['formSII'] = data['formSII'].save(commit=False)
                        if data['formAc'].is_valid():
                            print(type(propiedad.SII.appraisal_certificate.pk))
                            Ac = DocumentAc.objects.get(pk = propiedad.SII.appraisal_certificate.pk)
                            if request.FILES.get('ArDocAc') != None:
                                Ac.archive = request.FILES.get('ArDocAc')
                            Ac.comment = request.POST['CDocAc']
                            Ac.type = 'CA'
                            Ac.save()
                            data['formSII'].appraisal_certificate = Ac
                        if data['formEs'].is_valid():
                            DB = DocumentDB.objects.get(pk = propiedad.SII.debt_certificate.pk)
                            if request.FILES.get('ArDocEs') != None:
                                DB.archive = request.FILES.get('ArDocEs')
                            DB.comment = request.POST['CDocEs']
                            DB.type = 'CD'
                            DB.save()
                            data['formSII'].debt_certificate = DB
                        data['formSII'].save()
                        prop.SII = data['formSII']
                        prop.save()
                        valores = walk_object(acq_id)
                        prop.stats_acquisition.total = valores[0]
                        prop.stats_acquisition.complete = valores[1]
                        prop.save()
                    return JsonResponse({})
                
            if request.POST.get("commentacq") == None:
                pass
            else:
                x = Change_property.objects.create(id_property = acq_id, type = 'Acq', user = data['staff'], comment = request.POST.get("commentacq"))
                x.save()

        else:
            prop = AcquisitionForm(instance = propiedad)
            loca = LocationForm(instance = propiedad.location)
            # Arquitectura
            arqui = ArquitectureForm()
            docex = DocExForm()
            doccip = DocCipForm()
            doccn = DocCnForm()
            docblue = DocBlueForm()
            docbuild = DocBuildPForm()
            docmr = DocMRForm()
            # Internos
            inter = InternalForm()
            doctypec = DocTypeCForm()
            docother = DocOtherForm()
            # Notaria
            notary = NotaryForm()
            docwr = DocWRForm()
            docdc = DocDCForm()
            docph = DocPHForm()
            doces = DocEsForm()
            # SII
            sii = SII_recordForm()
            docac = DocAcForm()
            docdb = DocDBForm()

            if propiedad.arquitecture != None:
                arqui = ArquitectureForm(instance= propiedad.arquitecture)
                docex = DocExForm(instance= propiedad.arquitecture.expropriation_mun)
                doccip = DocCipForm(instance=propiedad.arquitecture.cip)
                doccn = DocCnForm(instance=propiedad.arquitecture.certified_number)
                docblue = DocBlueForm(instance=propiedad.arquitecture.blueprints)
                docbuild = DocBuildPForm(instance=propiedad.arquitecture.building_permit)
                docmr = DocMRForm(instance=propiedad.arquitecture.municipal_reception)
            if propiedad.internal != None:
                inter = InternalForm(instance=propiedad.internal)
                doctypec = DocTypeCForm(instance = propiedad.internal.contract_type)
                docother = DocOtherForm(instance=propiedad.internal.others)
            if propiedad.notary != None:
                notary = NotaryForm(instance=propiedad.notary)
                docwr = DocWRForm(instance=propiedad.notary.writing)
                docdc = DocDCForm(instance=propiedad.notary.domain_certificate)
                docph = DocPHForm(instance=propiedad.notary.prohibitions)
                doces = DocEsForm(instance=propiedad.notary.expropriation_serviu)
            if propiedad.SII != None:
                sii = SII_recordForm(instance=propiedad.SII)
                docac = DocAcForm(instance=propiedad.SII.appraisal_certificate)
                docdb = DocDBForm(instance=propiedad.SII.debt_certificate)
            data = {
                'formLocation': loca, 'formAcquisition': prop, 'formArquitecture':arqui,
                'formEx': docex, 'formCip': doccip, 'formCn': doccn,
                'formBlue': docblue,
                'formBuildP': docbuild, 'formMR': docmr, 'formInternal': inter,
                'formTypeC': doctypec,
                'formOther': docother, 'formNotary': notary, 'formWR': docwr,
                'formDC': docdc,
                'formPH': docph, 'formEs': doces, 'formSII': sii,
                'formAc': docac, 'formDB': docdb, 'id': acq_id
            }
        return render(request, 'edit_acquisition.html', data)
    else:
        return HttpResponseRedirect(reverse('list_acquisition'))


@login_required(login_url='login')
def Edit_rent(request, rent_id):
    data = {}
    print(rent_id)
    propiedad = get_object_or_404(Rent, pk=rent_id)
    data['staff'] = Staff.objects.get(username_staff=request.user)
    if data['staff'].type_user == 'DIG' or data['staff'].type_user == 'ADM':
        total = 0
        contestado = 0
        if propiedad.contract_type.archive != '':
            contestado += 1
        if propiedad.image != '':
            contestado +=1
        if request.POST:
            FormRent = RentForm(request.POST, request.FILES, instance=Rent.objects.get(pk=rent_id))
            FormLocation = LocationForm(request.POST, request.FILES,instance=propiedad.location)
            FormDocTypeC = DocTypeCForm(request.POST, request.FILES, instance=propiedad.contract_type)
            if FormRent.is_valid():
                if FormLocation.is_valid():
                    if FormDocTypeC.is_valid():
                        FormRent.save(commit=False)
                        FormLocation.save(commit=False)
                        FormDocTypeC.save(commit=False)
                        FormRent.save()
                        create_notification(request.user, "ER", rent_id, 'RT')
                        FormLocation.save()
                        FormDocTypeC.save()
                        if request.POST['plot'] == '':
                            print('vacio')
                            total -= 1
                        if request.POST['lot_number'] == '':
                            print('vacio2')
                            total -= 1
                        for val in request.POST:
                            print(val, request.POST[val])
                            if request.POST[val] != '':
                                contestado += 1
                                total += 1
                            else:
                                total += 1
                        print(propiedad.stats_rent)
                        stats = Stats.objects.get(pk = propiedad.stats_rent.pk)
                        stats.total = total-2
                        stats.complete = contestado-2
                        stats.save()
                        x = Change_property.objects.create(id_property = rent_id, type = 'Rent', user = data['staff'], comment = request.POST.get("commentrent"))
                        x.save()
                return redirect('list_rent')
        data['data'] = RentForm(instance=Rent.objects.get(pk=rent_id))
        data['pk'] = rent_id
        data['FormRent']  = RentForm(instance=propiedad)
        data['FormLocation']= LocationForm(instance=propiedad.location)
        print(propiedad.contract_type)
        data['FormDocTypeC'] = DocTypeCForm(instance=propiedad.contract_type)
        template_name = 'edit_rent.html'
        return render(request, template_name, data)
    else:
        return HttpResponseRedirect(reverse('list_acquisition'))

@login_required(login_url='login')
def Delete_rent(request,id):
    data = {}
    data['rent'] = Rent.objects.all()
    Rent.objects.filter(pk=id).delete()
    return HttpResponseRedirect(reverse('list_rent'))


@login_required(login_url='login')
def view_archive(request, document_type, document_id):
    data = {}
    data['request'] = request
    template_name = 'view_archive.html'
    if document_type == 'EM':
        if document_id != None:
            data['archive'] = DocumentEx.objects.get(pk=document_id)
            return render(request, template_name, data)
        else:
            return render(request, template_name)
    elif document_type == 'CP':
        if document_id != None:
            data['archive'] = DocumentCip.objects.get(pk=document_id)
            return render(request, template_name, data)
        else:
            return render(request, template_name)
    elif document_type == 'NC':
        if document_id != None:
            data['archive'] = DocumentCn.objects.get(pk=document_id)
            return render(request, template_name, data)
        else:
            return render(request, template_name)
    elif document_type == 'PL':
        if document_id != None:
            data['archive'] = DocumentBlue.objects.get(pk=document_id)
            return render(request, template_name, data)
        else:
            return render(request, template_name)
    elif document_type == 'PE':
        if document_id != None:
            data['archive'] = DocumentBuildP.objects.get(pk=document_id)
            return render(request, template_name, data)
        else:
            return render(request, template_name)
    elif document_type == 'RM':
        if document_id != None:
            data['archive'] = DocumentMR.objects.get(pk=document_id)
            return render(request, template_name, data)
        else:
            return render(request, template_name)
    elif document_type == 'TC':
        if document_id != None:
            data['archive'] = DocumentTypeC.objects.get(pk=document_id)
            return render(request, template_name, data)
        else:
            return render(request, template_name)
    elif document_type == 'OT':
        if document_id != None:
            data['archive'] = DocumentOther.objects.get(pk=document_id)
            return render(request, template_name, data)
        else:
            return render(request, template_name)
    elif document_type == 'ES':
        if document_id != None:
            data['archive'] = DocumentWR.objects.get(pk=document_id)
            return render(request, template_name, data)
        else:
            return render(request, template_name)
    elif document_type == 'DC':
        if document_id != None:
            data['archive'] = DocumentDC.objects.get(pk=document_id)
            return render(request, template_name, data)
        else:
            return render(request, template_name)
    elif document_type == 'SE':
        if document_id != None:
            data['archive'] = DocumentEs.objects.get(pk=document_id)
            return render(request, template_name, data)
        else:
            return render(request, template_name)
    elif document_type == 'CA':
        if document_id != None:
            data['archive'] = DocumentAc.objects.get(pk=document_id)
            return render(request, template_name, data)
        else:
            return render(request, template_name)
    elif document_type == 'CD':
        if document_id != None:
            data['archive'] = DocumentDB.objects.get(pk=document_id)
            return render(request, template_name, data)
        else:
            return render(request, template_name)
    elif document_type == 'PR':
        if document_id != None:
            data['archive'] = DocumentPH.objects.get(pk=document_id)
            return render(request, template_name, data)
        else:
            return render(request, template_name)
    else:
        return render(request, template_name)


@login_required(login_url='login')
def list_total(request):
    data = {}
    object_list_acquisition = Acquisition.objects.all().order_by('id')
    object_list_rent = Rent.objects.all().order_by('id')

    paginator_acq = Paginator(object_list_acquisition, 5)
    paginator_rent = Paginator(object_list_rent, 5)
    page = request.GET.get('page')
    data['staff'] = Staff.objects.get(username_staff = request.user)
    try:
        data['object_list_acquisition'] = paginator_acq.page(page)
        data['object_list_rent'] = paginator_rent.page(page)
    except PageNotAnInteger:
        data['object_list_acquisition'] = paginator_acq.page(1)
        data['object_list_rent'] = paginator_rent.page(1)
    except EmptyPage:
        data['object_list_acquisition'] = paginator_acq.page(paginator_acq.num_pages)
        data['object_list_rent'] = paginator_rent.page(paginator_rent.num_pages)
    template_name = 'list_total.html'
    return render(request, template_name, data)

@login_required(login_url='login')
def search(request):
    template = 'search.html'
    data = {}
    data[request] = request
    if request.method == "GET":
        selection = request.GET.get('selection')
        if selection == '1':
            query = request.GET.get('q')
            object_list_acquisition = Acquisition.objects.filter(Q(property_use=query) | Q(name__contains=query)).order_by('id')
            object_list_rent = Rent.objects.filter(Q(property_use=query) | Q(name__contains=query)).order_by('id')
            paginator_acq = Paginator(object_list_acquisition, 50)
            paginator_rent = Paginator(object_list_rent, 50)
            page = request.GET.get('page')
            try:
                data['object_list_acquisition'] = paginator_acq.page(page)
                data['object_list_rent'] = paginator_rent.page(page)
            except PageNotAnInteger:
                data['object_list_acquisition'] = paginator_acq.page(1)
                data['object_list_rent'] = paginator_rent.page(1)
            except EmptyPage:
                data['object_list_acquisition'] = paginator_acq.page(paginator_acq.num_pages)
                data['object_list_rent'] = paginator_rent.page(paginator_rent.num_pages)
            return render(request, template, data)
        elif selection == '2':
            query = request.GET.get('q')
            object_list_acquisition = Acquisition.objects.filter(
                Q(property_use=query) | Q(name__contains=query)).order_by('id')
            paginator_acq = Paginator(object_list_acquisition, 100)
            page = request.GET.get('page')
            try:
                data['object_list_acquisition'] = paginator_acq.page(page)
            except PageNotAnInteger:
                data['object_list_acquisition'] = paginator_acq.page(1)
            except EmptyPage:
                data['object_list_acquisition'] = paginator_acq.page(paginator_acq.num_pages)
            return render(request, template, data)
        elif selection == '3':
            query = request.GET.get('q')
            object_list_rent = Rent.objects.filter(Q(property_use__name=str(query)) | Q(name__contains=str(query)))
            paginator_rent = Paginator(object_list_rent, 100)
            page = request.GET.get('page')
            try:
                data['object_list_rent'] = paginator_rent.page(page)
            except PageNotAnInteger:
                data['object_list_rent'] = paginator_rent.page(1)
            except EmptyPage:
                data['object_list_rent'] = paginator_rent.page(paginator_rent.num_pages)
            return render(request, template, data)
        else:
            return render(request, 'list_total.html')
    else:
        return render(request, 'list_total.html')


def createstaff(request):
    data = {}
    data['title'] = 'Agregar Personal'
    data['staff'] = Staff.objects.get(username_staff=request.user)
    if data['staff'].type_user == 'ADM':
        print('primer if')
        if request.method == 'POST':
            print('segundo if')
            data['form'] = StaffForm(request.POST)
            data['form2'] = UserForm(request.POST)
            print(request.POST['password1'])
            print(request.POST['password2'])
            if data['form2'].is_valid():
                print('tercer if')
                if data['form'].is_valid():
                    print('cuarto if')
                    sav = data['form'].save(commit=False)
                    password = request.POST['password1']
                    print(password)
                    password2 = request.POST['password2']
                    print(password2)
                    username = request.POST['username']
                    print(username)
                    if User.objects.filter(username = username).exists() == False:
                        if User.objects.filter(email = request.POST['email']).exists() == False:
                            if password == password2:
                                print('quinto if')
                                sav2 = User.objects.create_user(username=request.POST['username'], password=request.POST['password1'], first_name=request.POST['first_name'], last_name=request.POST['last_name'], email=request.POST['email'],)
                                sav.username_staff = sav2
                                sav.save()
                                return HttpResponseRedirect(reverse('createstaff'))
                            else:
                                print('primer else')
                                data['resultado'] = 'Las contraseñas son diferentes'
                                return render(request, 'createstaff.html', data)
                        else:
                            data['resultado'] = 'El email ya esta registrado'
                            return render(request, 'createstaff.html', data)
                    else:
                        data['resultado'] = 'El nombre de usuario ya esta registrado'
                        return render(request, 'createstaff.html', data)
        else:
            print('Segundo else')
            data['form2'] = UserForm()
            data['form'] = StaffForm()
            template = 'createstaff.html'
            return render(request, template, data)
    else:
        print('Tercer else')
        return HttpResponseRedirect(reverse('list_total'))
        template = 'createstaff.html'

        return render(request, template, data)

def list_staff(request):
    data = {}
    staff = Staff.objects.get(username_staff = request.user)
    data['staff'] = staff
    if staff.type_user == 'ADM':
        object_list = Staff.objects.all().order_by('-id')
        paginator = Paginator(object_list, 10)
        page = request.GET.get('page')
        users = User.objects.all()
        try:
            data['object_list'] = paginator.page(page)
        except PageNotAnInteger:
        # If page is not an integadder, deliver first page.
            data['object_list'] = paginator.page(1)
        except EmptyPage:
            data['object_list'] = paginator.page(paginator.num_pages)
        template_name = 'list_staff.html'
        return render(request, template_name, data)
    else:
        return render(request, 'index.html', data)

def edit_staff(request,staff_id):
    data = {}
    staff = Staff.objects.get(username_staff = request.user)
    data['staff'] = staff
    staff = get_object_or_404(Staff, pk=staff_id)
    if request.method == 'POST':
        print(request.POST)
        form = EditStaffForm(request.POST, instance=Staff.objects.get(pk=staff_id))
        if form.is_valid():
            print('isvalid')
            if request.POST['status'] == 'True':
                print('True')
                staff.status = True
            elif request.POST['status'] == 'False':
                print('False')
                staff.status = False
            user = User.objects.get(pk = request.POST['pk'])
            user.first_name = request.POST['first_name']
            user.last_name = request.POST['last_name']
            user.email = request.POST['email']
            staff.save()
            user.save()
            form.save()
            return redirect('list_staff')
        else:
            data['data'] = EditStaffForm(instance=Staff.objects.get(pk=staff_id))
            user = User.objects.get(username=staff.username_staff)
            print(staff.type_user)
            data['user'] = user.pk
            return render(request, 'edit_staff.html', data)
    else:
        data['data'] = EditStaffForm(instance=Staff.objects.get(pk=staff_id))
        user = User.objects.get(username = staff.username_staff)
        print(staff.type_user)
        data['user'] = user.pk
        return render(request, 'edit_staff.html', data)

def edit_password(request, staff_id):
    data = {}
    staff = get_object_or_404(Staff, pk=staff_id)
    if request.method == 'POST':
        print(request.POST['status'])
        user = User.objects.get(pk=request.POST['pk'])
        old_password = request.POST.get('old_password')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        if user.check_password(old_password) == True:
            if password1 == password2:
                user.set_password(password1)
                user.save()
                return JsonResponse({'result':True})
            else:
                return JsonResponse({'result': 'difpass'})
        else:
            return JsonResponse({'result': 'incpass'})
    else:
        data['data'] = PasswordChangeForm(instance=User.objects.get(username = staff.username_staff))
        user = User.objects.get(username=staff.username_staff)
        data['user'] = user.pk
        return render(request, 'edit_password.html', data)

def edit_password(request, staff_id):
    data = {}
    staff = get_object_or_404(Staff, pk=staff_id)
    log = Staff.objects.get(pk=request.user.pk)
    logu = User.objects.get(pk=request.user.pk)
    if request.method == 'POST':
        user = User.objects.get(pk=request.POST.get('pk'))
        old_password = request.POST.get('old_password')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        if log.type_user == 'ADM':
            if logu.check_password(old_password) == True:
                if password1 == password2:
                    print('adm 3')
                    user.set_password(password1)
                    user.save()
                    return JsonResponse({'result': 'True'})
                else:
                    return JsonResponse({'result': 'difpass'})
            else:
                return JsonResponse({'result': 'incpassadm'})
        else:
            if user.check_password(old_password) == True:
                if password1 == password2:
                    user.set_password(password1)
                    user.save()
                    return JsonResponse({'result':True})
                else:
                    return JsonResponse({'result': 'difpass'})
            else:
                return JsonResponse({'result': 'incpassuser'})
    else:
        data['data'] = PasswordChangeForm(instance=User.objects.get(username = staff.username_staff))
        user = User.objects.get(username=staff.username_staff)
        data['user'] = user.pk
        data['staff'] = log
        return render(request, 'edit_password.html', data)
def Email_password(request):
    template = 'email_password.html'
    data = {}
    if request.POST:
        print (request.POST)
        print('x')
        print(User.objects.filter(email=request.POST['email']).exists())
        if User.objects.filter(email=request.POST['email']).exists() == True:
            users = User.objects.get(email = request.POST['email'])
            print(users)
            if request.POST['email'] == str(users.email):

                randomString()
                send_mail(
                    'Recuperar contraseña',
                    'Su codigo es: ' + CODE + '',
                    'ezticketdev@gmail.com',
                    [request.POST['email']],
                    fail_silently=False
                )
                global user_pass
                user_pass = users
                return JsonResponse({'result':True})
        else:
            return JsonResponse({'result':False})

    return render(request, template, data)

def Forgot_password(request):
    template = 'forgot_password.html'
    data = {}
    if request.POST:
        user_profile =User.objects.get(username=user_pass)
        user_profile.set_password(request.POST['pass'])
        user_profile.save()
        return JsonResponse({'result':True})
    return render(request, template, data)

def Validate(request):
    template = 'validate.html'
    data = {}
    if request.POST:
        user_profile =Userprofile.objects.get(user=request.user)
        if CODE == request.POST['code']:
            user_profile.verificate = True
            user_profile.save()
            return JsonResponse({'result':True})
        return JsonResponse({'result':False})

    return render(request, template, data)

def Validate_password(request):
    template = 'validate_password.html'
    data = {}
    if request.POST:
        if CODE == request.POST['code']:
            return JsonResponse({'result':True})
        return JsonResponse({'result':False})
    return render(request, template, data)

def add_region(request):
    template = 'add_region.html'
    data = {}
    data['staff'] = Staff.objects.get(username_staff = request.user)
    print(request.POST.get('acronym'))
    p = str(request.POST.get('name'))
    q = str(request.POST.get('acronym'))
    p = p.upper()
    q = q.upper()
    if Region.objects.filter(name=p).exists() == False:
        if len(q)<=5:
            if request.POST:
                x = Region.objects.create(name=p, acronym=q)
                x.save()
                return JsonResponse({'result':True})
        else:
            return JsonResponse({'result': 'Max'})
    else:
        return JsonResponse({'result':False})

    return render(request, template, data)

def edit_region(request, region_id):
    data = {}
    template = 'edit_region.html'
    data['info'] = Region.objects.get(pk = region_id)
    if request.method == 'POST':
        nombre = request.POST.get('name')
        acronimo = request.POST.get('acronym')
        Region.objects.filter(pk = region_id).update(name = nombre, acronym = acronimo)

        return JsonResponse({'result': True})
    else:
        return render(request, template, data)

def delete_region(request, region_id):
    template = 'delete_region.html'
    data = {}
    data['info'] = Region.objects.get(pk = region_id)
    if request.method == 'POST':
        if request.POST['valor'] == 'True':
            Region.objects.filter(pk = region_id).delete()
            return JsonResponse({'result': True})
        else:
            return JsonResponse({'result': False})
    else:
        return render(request, template, data)

def list_region(request):
    template = 'list_region.html'
    data = {}
    staff = Staff.objects.get(username_staff = request.user)
    data['staff'] = staff
    object_list = Region.objects.all().order_by()
    paginator = Paginator(object_list, 20)
    page = request.GET.get('page')
    try:
        data['object_list'] = paginator.page(page)
    except PageNotAnInteger:
        data['object_list'] = paginator.page(1)
    except EmptyPage:
        data['object_list'] = paginator.page(paginator.num_pages)
    return render(request, template, data)

def change_status_acquisition(request, id):
    aux = Acquisition.objects.get(pk = id)
    if aux.status == True:
        Acquisition.objects.filter(pk = id).update(status = False)
        create_notification(request.user, "SI", id, 'AC')
    elif aux.status == False:
        Acquisition.objects.filter(pk = id).update(status = True)
        create_notification(request.user, "SA", id, 'AC')
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def change_status_rent(request, id):
    aux = Rent.objects.get(pk = id)
    if aux.status == True:
        Rent.objects.filter(pk = id).update(status = False)
        create_notification(request.user, "SI", id, 'RT')
    elif aux.status == False:
        Rent.objects.filter(pk = id).update(status = True)
        create_notification(request.user, "SA", id, 'RT')
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def add_property(request):
    template = 'add_property.html'
    data = {}
    data['staff'] = Staff.objects.get(username_staff = request.user)
    p = str(request.POST.get('name'))
    q = str(request.POST.get('acronym'))
    p = p.upper()
    q = q.upper()
    if Region.objects.filter(name=p).exists() == False:
        if len(q)<=5:
            if request.POST:
                x = Property.objects.create(name=p, acronym=q)
                x.save()
                return JsonResponse({'result':True})
        else:
            return JsonResponse({'result': 'Max'})
    else:
        return JsonResponse({'result':False})

    return render(request, template, data)

def edit_property(request, property_id):
    data = {}
    template = 'edit_property.html'
    data['info'] = Property.objects.get(pk = property_id)
    x = Property.objects.get(pk = property_id)
    data['version_list'] = nn = Version.objects.get_for_object(x)
    if request.method == 'POST':
        nombre = request.POST.get('name')
        acronimo = request.POST.get('acronym')
        if x.name != nombre:
            x.name = nombre
            if x.acronym != acronimo:
                x.acronym = acronimo
                x.save()
            else:
                x.save()
        else:
            if x.acronym != acronimo:
                x.acronym = acronimo
                x.save()
            else:
                x.save()

        return JsonResponse({'result': True})
    else:
        return render(request, template, data)

def delete_property(request, property_id):
    template = 'delete_property.html'
    data = {}
    data['info'] = Property.objects.get(pk = property_id)
    if request.method == 'POST':
        if request.POST['valor'] == 'True':
            Property.objects.filter(pk = property_id).delete()
            return JsonResponse({'result': True})
        else:
            return JsonResponse({'result': False})
    else:
        return render(request, template, data)

def list_property(request):
    template = 'list_property.html'
    data = {}
    staff = Staff.objects.get(username_staff = request.user)
    data['staff'] = staff
    object_list = Property.objects.all().order_by()
    paginator = Paginator(object_list, 20)
    page = request.GET.get('page')
    try:
        data['object_list'] = paginator.page(page)
    except PageNotAnInteger:
        data['object_list'] = paginator.page(1)
    except EmptyPage:
        data['object_list'] = paginator.page(paginator.num_pages)
    return render(request, template, data)

def create_staff(request):
    data = {}
    data['title'] = 'Agregar Personal'
    data['staff'] = Staff.objects.get(username_staff=request.user)
    if data['staff'].type_user == 'ADM':
        if request.method == 'POST':
            data['form'] = StaffForm(request.POST)
            data['form2'] = UserForm(request.POST)
            if data['form2'].is_valid():
                if data['form'].is_valid():
                    sav = data['form'].save(commit=False)
                    password = request.POST.get('password1')
                    password2 = request.POST.get('password2')
                    print(password_valid(password))
                    if password_valid(password) == True:
                        if password == password2:
                            sav2 = User.objects.create_user(username=request.POST.get('username'), password=request.POST.get('password1'), first_name=request.POST.get('first_name'), last_name=request.POST.get('last_name'), email=request.POST.get('email'))
                            sav.username_staff = sav2
                            sav.save()
                            return JsonResponse({'result': 'True'})
                        else:
                            return JsonResponse({'result': 'password'})
                    else:
                        return JsonResponse({'result': 'valid'})

                else:
                    data['form2'] = UserForm()
                    data['form'] = StaffForm()
                    template = 'createstaff.html'
                    return render(request, template, data)
            else:
                if User.objects.filter(username = request.POST.get('username')).exists():
                    if User.objects.filter(email = request.POST.get('email')).exists():
                        return JsonResponse({'result':'email'})
                    else:
                        return JsonResponse({'result': 'username'})
        else:
            data['form2'] = UserForm()
            data['form'] = StaffForm()
            template = 'createstaff.html'
            return render(request, template, data)
    else:
        return HttpResponseRedirect(reverse('list_total'))
    template = 'createstaff.html'
    return render(request, template, data)

#Valida que la contraseña cumpla con las caracteristicas minimas
def password_valid(password):
    flag = 0
    while True:
        if (len(password)<8):
            flag = -1
            break
        elif not re.search("[a-z]", password):
            flag = -1
            break
        elif not re.search("[A-Z]", password):
            flag = -1
            break
        elif not re.search("[0-9]", password):
            flag = -1
            break
        else:
            flag = 0
            print("Valid Password")
            return True
            break

    if flag ==-1:
        print("Not a Valid Password")
        return False

# Post de usuarios

def add_post(request):
    data = {}
    if request.method == "POST" and request.POST.get('description') != '':
        author_user = Staff.objects.get(username_staff=request.user)
        description_post = str(request.POST.get('description'))
        if request.POST.get('prop') == 'acquisition':
            id = request.POST.get('prop_id')
            acquisition_post = Acquisition.objects.get(pk=id)
            n_post = Post.objects.create(
                author = author_user,
                description = description_post,
                acquisition = acquisition_post,
            )
            n_post.save()
            create_notification(request.user, "PA", id, 'AC')
            data = {'result': True}
            return JsonResponse(data)
        elif request.POST.get('prop') == 'rent' and request.POST.get('description') != '':
            id = request.POST.get('prop_id')
            rent_post = Rent.objects.get(pk=id)
            n_post = Post.objects.create(
                author=author_user,
                description=description_post,
                rent=rent_post,
            )
            n_post.save()
            create_notification(request.user, "PR", id, 'RT')
            data = {'result': True}
            return JsonResponse(data)
        else:
            data = {'result': False}
            return JsonResponse(data)

    else:
        data = {'result': False}
        return JsonResponse(data)

def edit_post(request):
    data = {}
    post_id = request.POST.get('id')
    print(request.POST.get('description'))
    post = Post.objects.get(pk = post_id)
    if request.method == 'POST':
        descrip = request.POST.get('description')
        Post.objects.filter(pk = post_id).update(description = descrip)

        return JsonResponse({'result': True})
    else:
        data = {'result': False}
        return JsonResponse(data)

def delete_post(request):
    data = {}
    if request.POST.get('id') is not None:
        if request.method == 'POST':
            id = request.POST.get('id')
            Post.objects.filter(pk = id).delete()
            return JsonResponse({'result': True})
        else:
            JsonResponse({'result': False})
    else:
        JsonResponse({'result': False})

def add_comment(request):
    data = {}
    print(request.POST)
    if request.method == "POST" and request.POST.get('description') != '':
        author_user = Staff.objects.get(username_staff=request.user)
        description_post = str(request.POST.get('description'))
        if request.POST.get('prop') == 'acquisition':
            id = request.POST.get('prop_id')
            post_id = Post.objects.get(pk=id)
            n_post = Comment.objects.create(
                author = author_user,
                description = description_post,
                post = post_id,
            )
            n_post.save()
            create_notification(request.user, "RA", id, 'AC')
            data = {'result': True}
            return JsonResponse(data)
        elif request.POST.get('prop') == 'rent':
            id = request.POST.get('prop_id')
            rent_post = Post.objects.get(pk=id)
            n_post = Comment.objects.create(
                author=author_user,
                description=description_post,
                post = rent_post,
            )
            n_post.save()
            create_notification(request.user, "RR", id, 'RT')
            data = {'result': True}
            return JsonResponse(data)
        else:
            data = {'result': False}
            return JsonResponse(data)

    else:
        data = {'result': False}
        return JsonResponse(data)

def edit_comment(request):
    data = {}
    comment_id = request.POST.get('id')
    post = Comment.objects.get(pk=comment_id)
    if request.method == 'POST':
        descrip = request.POST.get('description')
        Comment.objects.filter(pk=comment_id).update(description = descrip)

        return JsonResponse({'result': True})
    else:
        data = {'result': False}
        return JsonResponse(data)

def delete_comment(request):
    data = {}
    if request.POST.get('id') is not None:
        if request.method == 'POST':
            id = request.POST.get('id')
            Comment.objects.filter(pk=id).delete()
            return JsonResponse({'result': True})
        else:
            JsonResponse({'result': False})
    else:
        JsonResponse({'result': False})

def create_stats(total, complete):
    complete = complete
    total = total
    percentage = (complete/total)*100
    x = Stats.objects.create(total = total, complete = complete, percentage = percentage)
    return x

def contador(x):
    total = 0
    nocontestado = 0
    for val in x:
        print(val, x[val])
        if x[val] != '':
            nocontestado += 1
            total += 1
        else:
            total += 1
    return (total, nocontestado)

def add_district(request):
    template = 'add_district.html'
    data = {}
    data['staff'] = Staff.objects.get(username_staff = request.user)
    p = str(request.POST.get('name'))
    q = str(request.POST.get('acronym'))
    print('---------------------')
    print(len(q))
    print('---------------------')
    p = p.upper()
    q = q.upper()
    if District.objects.filter(name=p).exists() == False:
        if len(q)<=5:
            if request.POST:
                x = District.objects.create(name=p, acronym=q)
                x.save()
                return JsonResponse({'result': 'True'})
        else:
            return JsonResponse({'result': 'Max'})
    else:
        return JsonResponse({'result':False})

    return render(request, template, data)

def edit_district(request, district_id):
    data = {}
    data['staff'] = Staff.objects.get(username_staff = request.user)
    template = 'edit_district.html'
    data['info'] = District.objects.get(pk = district_id)
    print(data)
    if request.method == 'POST':
        nombre = request.POST.get('name')
        acronimo = request.POST.get('acronym')
        District.objects.filter(pk = district_id).update(name = nombre, acronym = acronimo)

        return JsonResponse({'result': True})
    else:
        return render(request, template, data)

def delete_district(request, district_id):
    template = 'delete_district.html'
    data = {}
    data['info'] = District.objects.get(pk = district_id)
    if request.method == 'POST':
        if request.POST['valor'] == 'True':
            District.objects.filter(pk = district_id).delete()
            return JsonResponse({'result': True})
        else:
            return JsonResponse({'result': False})
    else:
        return render(request, template, data)

def list_district(request):
    template = 'list_district.html'
    data = {}
    data['staff'] = Staff.objects.get(username_staff = request.user)
    object_list = District.objects.all().order_by()
    paginator = Paginator(object_list, 20)
    page = request.GET.get('page')
    try:
        data['object_list'] = paginator.page(page)
    except PageNotAnInteger:
        data['object_list'] = paginator.page(1)
    except EmptyPage:
        data['object_list'] = paginator.page(paginator.num_pages)
    return render(request, template, data)

def generate_report_excel(request):
    # para mandar al html que es un Excel
    if request.POST.get('texto') != None:
        query = request.POST.get('texto')
        if request.POST.get('filtro') != None:
            if request.POST.get('filtro') == 'comuna_true':
                object_list_acquisition = Acquisition.objects.filter(
                    Q(location__commune=query)).order_by('id')
            if request.POST.get('filtro') == 'ciudad_true':
                object_list_acquisition = Acquisition.objects.filter(
                    Q(location__city=query)).order_by('id')
            if request.POST.get('filtro') == 'region_true':
                object_list_acquisition = Acquisition.objects.filter(
                    Q(location__region__name=query)).order_by('id')
            if request.POST.get('filtro') == 'tipo_true':
                object_list_acquisition = Acquisition.objects.filter(
                    Q(writing_data=query)).order_by('id')
            if request.POST.get('filtro') == 'uso_true':
                object_list_acquisition = Acquisition.objects.filter(
                    Q(property_use__name=query)).order_by('id')
    else:
        return JsonResponse({'result': False})

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Reporte_propiedades.xlsx'
    excel = Report()
    time = str(timezone.now())
    datetime = str(time)
    excel.date = time
    filename_excel = 'Reporte_propiedades' + datetime + '.xlsx'
    
    # Hoja de trabajo
    wb = Workbook()
    ws = wb.active
    ft = Font(bold=True)
    bor = Border(left=Side(border_style='thin',color = 'FF000000'),
                    right = Side(border_style='thin',color = 'FF000000'),
                    top = Side(border_style='thin',color = 'FF000000'),
                    bottom = Side(border_style='thin',color = 'FF000000'))
    align = Alignment(wrap_text=True,vertical='center',horizontal='center')
    ws['C2'] = 'Organización Adventista del Séptimo Día'
    ws['C2'].font = ft
    ws['C3'] = 'Reporte propiedades'
    ws['C3'].font = ft
    time = str(timezone.now())
    ws['C4'] = str(time[:10])
    ws.merge_cells('C2:F2')
    ws.merge_cells('C3:F3')
    img = Image('static/logoA.png')
    img.width = 50
    img.height = 50
    ws.add_image(img,'B2')
    my_style = NamedStyle(name='Encabezado_Tabla')
    my_style.font = ft
    my_style.border = bor
    my_style.alignment = align
    my_style.fill = PatternFill("solid", fgColor="DDDDDD")

    my_style2 = NamedStyle(name='Contenido_Tabla')
    my_style2.border = bor
    my_style2.alignment = align

    #cabeceras de la tabla
    ws['B6'] = 'N°AASI'
    ws['B6'].style = my_style

    ws['C6'] = 'N°ROL'
    ws['C6'].style = my_style

    ws['D6'] = 'Uso'
    ws['D6'].style = my_style

    ws['E6'] = 'Nombre'
    ws['E6'].style = my_style
    ws.merge_cells('E6:G6')

    ws['H6'] = 'Dirección'
    ws['H6'].style = my_style
    ws.merge_cells('H6:J6')

    ws['K6'] = 'Nro.'
    ws['K6'].style = my_style

    ws['L6'] = 'Comuna'
    ws['L6'].style = my_style
    ws.merge_cells('L6:M6')

    ws['N6'] = 'Ciudad'
    ws['N6'].style = my_style

    ws['O6'] = 'Región'
    ws['O6'].style = my_style

    ws['P6'] = 'Escritura'
    ws['P6'].style = my_style

    ws['Q6'] = 'Fecha inscripción'
    ws['Q6'].style = my_style
    ws.merge_cells('Q6:R6')

    ws['S6'] = 'Vendedor'
    ws['S6'].style = my_style
    ws.merge_cells('S6:U6')

    ws['V6'] = 'Comprador'
    ws['V6'].style = my_style
    ws.merge_cells('V6:X6')

    ws['Y6'] = 'Notaría'
    ws['Y6'].style = my_style
    ws.merge_cells('Y6:AB6')

    ws['AC6'] = 'Año Escritura'
    ws['AC6'].style = my_style

    ws['AD6'] = 'Título Anterior'
    ws['AD6'].style = my_style

    ws['AE6'] = 'Título Actual'
    ws['AE6'].style = my_style

    ws['AF6'] = 'Precio Venta'
    ws['AF6'].style = my_style
    ws.merge_cells('AF6:AG6')

    ws['AH6'] = 'Valor Terreno'
    ws['AH6'].style = my_style
    ws.merge_cells('AH6:AI6')

    ws['AJ6'] = 'Valor construcción'
    ws['AJ6'].style = my_style
    ws.merge_cells('AJ6:AK6')

    ws['AL6'] = 'Nombre Propietario SII'
    ws['AL6'].style = my_style
    ws.merge_cells('AL6:AN6')

    ws['AO6'] = 'Destino SII'
    ws['AO6'].style = my_style
    ws.merge_cells('AO6:AQ6')

    ws['AR6'] = 'Deuda total'
    ws['AR6'].style = my_style
    ws.merge_cells('AR6:AS6')

    ws['AT6'] = 'Avalúo fiscal'
    ws['AT6'].style = my_style
    ws.merge_cells('AT6:AV6')

    ws['AW6'] = 'Exención contribuciones'
    ws['AW6'].style = my_style
    ws.merge_cells('AW6:AX6')

    ws['AY6'] = 'Superficie Terreno'
    ws['AY6'].style = my_style
    ws.merge_cells('AY6:AZ6')

    ws['BA6'] = 'MTS construidos PE-RF'
    ws['BA6'].style = my_style
    ws.merge_cells('BA6:BB6')

    ws['BC6'] = 'MTS construidos existentes'
    ws['BC6'].style = my_style
    ws.merge_cells('BC6:BD6')

    ws['BE6'] = 'Nro Recepción municipal'
    ws['BE6'].style = my_style
    ws.merge_cells('BE6:BF6')

    ws['BG6'] = 'Nro permiso edificación'
    ws['BG6'].style = my_style
    ws.merge_cells('BG6:BH6')



    #LLamar a objetos
    ven = 'Sin registro'
    adq = 'Sin registro'
    nota = 'Sin registro'
    wy = 'Sin registro'
    pt = 'Sin registro'
    ct = 'Sin registro'
    sp = 'Sin registro'
    vl = 'Sin registro'
    vc = 'Sin registro'
    des = 'Sin registro'
    owner = 'Sin registro'
    dest = 'Sin registro'
    debt = 'Sin registro'
    af = 'Sin registro'
    ec = 'Sin registro'
    st = 'Sin registro'
    mc = 'Sin registro'
    me = 'Sin registro'
    rm = 'Sin registro'
    pe = 'Sin registro'
    
    for prop in object_list_acquisition:
        #Recorrer objetos
        cont = 7
        calle = prop.location.street
        if prop.location.plot != None and prop.location.lot_number != None:
            calle = str(calle) + ',' + str(prop.location.plot) + ' ' + str(prop.location.lot_number)

        if prop.internal is not None:
            if prop.internal.supplier_name is not None:
                ven = prop.internal.supplier_name
            if prop.internal.acquiring_name is not None:
                adq = prop.internal.acquiring_name
            if prop.internal.value_land is not None:
                vl = prop.internal.value_land
            if prop.internal.value_construction is not None:
                vc = prop.internal.value_construction

        if prop.notary is not None:
            if prop.notary.notary is not None:
                nota = prop.notary.notary
            if prop.notary.writing_year is not None:
                wy = prop.notary.writing_year
            if prop.notary.previous_title is not None:
                pt = prop.notary.previous_title
            if prop.notary.current_title is not None:
                ct = prop.notary.current_title
            if prop.notary.sale_price is not None:
                sp = prop.notary.sale_price

        if prop.SII is not None:
            if prop.SII.destiny is not None:
                des = prop.SII.destiny
            if prop.SII.owner_name_SII is not None:
                owner = prop.SII.owner_name_SII
            if prop.SII.total_debt is not None:
                debt = prop.SII.total_debt
            if prop.SII.tax_appraisal is not None:
                af = prop.SII.tax_appraisal
            if prop.SII.ex_contributions is not None:
                ec = prop.SII.ex_contributions

        if prop.arquitecture is not None:
            if prop.arquitecture.ground_surface is not None:
                st = prop.arquitecture.ground_surface
            if prop.arquitecture.square_m_build is not None:
                mc = prop.arquitecture.square_m_build
            if prop.arquitecture.e_construction_m is not None:
                me = prop.arquitecture.e_construction_m
            if prop.arquitecture.municipal_n is not None:
                rm = prop.arquitecture.municipal_n
            if prop.arquitecture.n_building_permit is not None:
                pe = prop.arquitecture.n_building_permit

        #escribir la tabla, en matrices
        ws.cell(row=cont, column=2).value = prop.number_AASI
        ws.cell(row=cont, column=2).style = my_style2

        ws.cell(row=cont, column=3).value = prop.role_number
        ws.cell(row=cont, column=3).style = my_style2

        ws.cell(row=cont,column=4).value = prop.property_use.acronym
        ws.cell(row=cont, column=4).style = my_style2

        ws.merge_cells(start_row=cont, start_column=5, end_row=cont, end_column=7)
        ws.cell(row=cont, column=5).value = prop.name
        ws.cell(row=cont, column=5).style = my_style2
        ws.cell(row=cont, column=6).style = my_style2
        ws.cell(row=cont, column=7).style = my_style2

        ws.merge_cells(start_row=cont, start_column=8, end_row=cont, end_column=10)
        ws.cell(row=cont, column=8).value = calle
        ws.cell(row=cont, column=8).style = my_style2
        ws.cell(row=cont, column=9).style = my_style2
        ws.cell(row=cont, column=10).style = my_style2

        ws.cell(row=cont, column=11).value = prop.location.number
        ws.cell(row=cont, column=11).style = my_style2

        ws.merge_cells(start_row=cont, start_column=12, end_row=cont, end_column=13)
        ws.cell(row=cont, column=12).value = prop.location.commune
        ws.cell(row=cont, column=12).style = my_style2
        ws.cell(row=cont, column=13).style = my_style2

        ws.cell(row=cont, column=14).value = prop.location.city
        ws.cell(row=cont, column=14).style = my_style2

        ws.cell(row=cont, column=15).value = prop.location.region.acronym
        ws.cell(row=cont, column=15).style = my_style2

        ws.cell(row=cont, column=16).value = prop.writing_data
        ws.cell(row=cont, column=16).style = my_style2

        ws.merge_cells(start_row=cont, start_column=17, end_row=cont, end_column=18)
        ws.cell(row=cont, column=17).value = str(prop.acquisition_date)[:10]
        ws.cell(row=cont, column=17).style = my_style2
        ws.cell(row=cont, column=18).style = my_style2

        ws.merge_cells(start_row=cont, start_column=19, end_row=cont, end_column=21)
        ws.cell(row=cont, column=19).value = ven
        ws.cell(row=cont, column=19).style = my_style2
        ws.cell(row=cont, column=20).style = my_style2
        ws.cell(row=cont, column=21).style = my_style2

        ws.merge_cells(start_row=cont, start_column=22, end_row=cont, end_column=24)
        ws.cell(row=cont, column=22).value = adq
        ws.cell(row=cont, column=22).style = my_style2
        ws.cell(row=cont, column=23).style = my_style2
        ws.cell(row=cont, column=24).style = my_style2

        ws.merge_cells(start_row=cont, start_column=25, end_row=cont, end_column=28)
        ws.cell(row=cont, column=25).value = nota
        ws.cell(row=cont, column=25).style = my_style2
        ws.cell(row=cont, column=26).style = my_style2
        ws.cell(row=cont, column=27).style = my_style2
        ws.cell(row=cont, column=28).style = my_style2

        ws.cell(row=cont, column=29).value = wy
        ws.cell(row=cont, column=29).style = my_style2

        ws.cell(row=cont, column=30).value = pt
        ws.cell(row=cont, column=30).style = my_style2

        ws.cell(row=cont, column=31).value = ct
        ws.cell(row=cont, column=31).style = my_style2

        ws.merge_cells(start_row=cont, start_column=32, end_row=cont, end_column=33)
        ws.cell(row=cont, column=32).value = sp
        ws.cell(row=cont, column=32).style = my_style2
        ws.cell(row=cont, column=33).style = my_style2

        ws.merge_cells(start_row=cont, start_column=34, end_row=cont, end_column=35)
        ws.cell(row=cont, column=34).value = vl
        ws.cell(row=cont, column=34).style = my_style2
        ws.cell(row=cont, column=35).style = my_style2

        ws.merge_cells(start_row=cont, start_column=36, end_row=cont, end_column=37)
        ws.cell(row=cont, column=36).value = vc
        ws.cell(row=cont, column=36).style = my_style2
        ws.cell(row=cont, column=37).style = my_style2

        ws.merge_cells(start_row=cont, start_column=38, end_row=cont, end_column=40)
        ws.cell(row=cont, column=38).value = owner
        ws.cell(row=cont, column=38).style = my_style2
        ws.cell(row=cont, column=39).style = my_style2
        ws.cell(row=cont, column=40).style = my_style2

        ws.merge_cells(start_row=cont, start_column=41, end_row=cont, end_column=43)
        ws.cell(row=cont, column=41).value = des
        ws.cell(row=cont, column=41).style = my_style2
        ws.cell(row=cont, column=42).style = my_style2
        ws.cell(row=cont, column=43).style = my_style2

        ws.merge_cells(start_row=cont, start_column=44, end_row=cont, end_column=45)
        ws.cell(row=cont, column=44).value = debt
        ws.cell(row=cont, column=44).style = my_style2
        ws.cell(row=cont, column=45).style = my_style2

        ws.merge_cells(start_row=cont, start_column=46, end_row=cont, end_column=48)
        ws.cell(row=cont, column=46).value = af
        ws.cell(row=cont, column=46).style = my_style2
        ws.cell(row=cont, column=47).style = my_style2
        ws.cell(row=cont, column=48).style = my_style2

        ws.merge_cells(start_row=cont, start_column=49, end_row=cont, end_column=50)
        ws.cell(row=cont, column=49).value = ec
        ws.cell(row=cont, column=49).style = my_style2
        ws.cell(row=cont, column=50).style = my_style2

        ws.merge_cells(start_row=cont, start_column=51, end_row=cont, end_column=52)
        ws.cell(row=cont, column=51).value = st
        ws.cell(row=cont, column=51).style = my_style2
        ws.cell(row=cont, column=52).style = my_style2

        ws.merge_cells(start_row=cont, start_column=53, end_row=cont, end_column=54)
        ws.cell(row=cont, column=53).value = mc
        ws.cell(row=cont, column=53).style = my_style2
        ws.cell(row=cont, column=54).style = my_style2

        ws.merge_cells(start_row=cont, start_column=55, end_row=cont, end_column=56)
        ws.cell(row=cont, column=55).value = me
        ws.cell(row=cont, column=55).style = my_style2
        ws.cell(row=cont, column=56).style = my_style2

        ws.merge_cells(start_row=cont, start_column=57, end_row=cont, end_column=58)
        ws.cell(row=cont, column=57).value = rm
        ws.cell(row=cont, column=57).style = my_style2
        ws.cell(row=cont, column=58).style = my_style2

        ws.merge_cells(start_row=cont, start_column=59, end_row=cont, end_column=60)
        ws.cell(row=cont, column=59).value = pe
        ws.cell(row=cont, column=59).style = my_style2
        ws.cell(row=cont, column=60).style = my_style2

        cont= cont+1
    
    wb.save('media/Reportes/'+filename_excel)

    print('se hace el excel')
    return JsonResponse({'url': '/media/Reportes/'+filename_excel})


def walk_object(id):
    prop = Acquisition.objects.get(pk=id)
    a = model_to_dict(prop)
    arq, internal, notary, SII = ArchitectureRecordAcq.objects.get(pk = a.get('arquitecture')), InternalAccountantsAcq.objects.get(pk = a.get('internal')), NotaryAcquisition.objects.get(pk = a.get('notary')), SiiRecord.objects.get(pk = a.get('SII'))
    d_arq, d_internal, d_notary, d_SII = model_to_dict(arq), model_to_dict(internal), model_to_dict(notary), model_to_dict(SII)
    #Elementos de arquitectura
    exp_num, cip, cer_num, bluprin, bui_per, mun_rec = DocumentEx.objects.get(pk = d_arq.get('expropriation_mun')), DocumentCip.objects.get(pk = d_arq.get('cip')), DocumentCn.objects.get(pk = d_arq.get('certified_number')), DocumentBlue.objects.get(pk = d_arq.get('blueprints')), DocumentBuildP.objects.get(pk = d_arq.get('building_permit')), DocumentMR.objects.get(pk = d_arq.get('municipal_reception'))
    #elementos de interno
    con_typ, other = DocumentTypeC.objects.get(pk = d_internal.get('contract_type')), DocumentOther.objects.get(pk = d_internal.get('others'))
    #elementos de notaria
    write, dom_cer, proh, exp_ser = DocumentWR.objects.get(pk=d_notary.get('writing')), DocumentDC.objects.get(pk = d_notary.get('domain_certificate')), DocumentPH.objects.get(pk=d_notary.get('prohibitions')), DocumentEs.objects.get(pk=d_notary.get('expropriation_serviu'))
    #elementos de SII
    app_cer, deb_cer = DocumentAc.objects.get(pk = d_SII.get('appraisal_certificate')), DocumentDB.objects.get(pk = d_SII.get('debt_certificate'))
    contestado = 0
    total = 0
    valores = []
    for val in d_arq:
        #print(val, d_arq[val])
        if d_arq[val] == None:
            total += 1
        else:
            total += 1
            contestado += 1
    if exp_num.archive != '':   
        total += 1
        contestado += 1
    else:
        total += 1
    #-------------------------
    if cip.archive != '':
        total += 1
        contestado += 1
    else:
        total += 1
    #------------------------
    if cer_num.archive != '':
        total += 1
        contestado += 1
    else:
        total += 1
    #------------------------
    if bluprin.archive != '':
        total += 1
        contestado += 1
    else:
        total += 1
    #------------------------
    if bui_per.archive != '':
        total += 1
        contestado += 1
    else:
        total += 1
    #------------------------
    if mun_rec.archive != '':
        total += 1
        contestado += 1
    else:
        total += 1
    #------------------------
    print('--------------------------------')
    for val in d_internal:
        #print(val, d_internal[val])
        if d_internal[val] == None:
            total += 1
        else:
            total += 1
            contestado += 1

    if con_typ.archive != '':
        total += 1
        contestado += 1
    else:
        total += 1
    #-------------------------
    if other.archive != '':
        total += 1
        contestado += 1
    else:
        total += 1
    #-------------------------
    print('-------------------------------')
    for val in d_notary:
        #print(val, d_notary[val])
        if d_notary[val] == None:
            total += 1
        else:
            total += 1
            contestado += 1

    if write.archive != '':
        total += 1
        contestado += 1
    else:
        total += 1
    #-------------------------
    if dom_cer.archive != '':
        total += 1
        contestado += 1
    else:
        total += 1
    #------------------------
    if proh.archive != '':
        total += 1
        contestado += 1
    else:
        total += 1
    #------------------------
    if exp_ser.archive != '':
        total += 1
        contestado += 1
    else:
        total += 1
    print('-------------------------')
    for val in d_SII:
        #print(val, d_SII[val])
        if d_SII[val] == None:
            total += 1
        else:
            total += 1
            contestado += 1
    if app_cer.archive != '':
        total += 1
        contestado += 1
    else:
        total += 1
    #-------------------------
    if deb_cer.archive != '':
        total += 1
        contestado += 1
    else:
        total += 1
    #-------------------------
    total -= 18
    contestado -=18
    valores.append(total)
    valores.append(contestado)
    return valores
def delete_notification(request):
    pk_noti = request.POST.get('id_notification')
    print(pk_noti)
    if request.POST.get('id_notification') != None:
        Notification.objects.filter(pk = pk_noti).delete()
        return JsonResponse({'result':True})
    else:
        return JsonResponse({'result':False})


def view_notificaction(request):
    data = {}
    data['Notifications'] = Notification.objects.filter(user_receiver__username_staff = request.user).order_by('-time')
    for noti in data['Notifications']:
        delta = timezone.now()-noti.time
        if delta.days != 0:
            Notification.objects.filter(pk = noti.pk).update(time_str = (str(delta.days)+' d'))
        else:
            aux = str(delta).split(':')
            if aux[0] != '0':
                Notification.objects.filter(pk = noti.pk).update(time_str=(aux[0]+' h'))
            else:
                if aux[1][0] != '0':
                    Notification.objects.filter(pk = noti.pk).update(time_str=(aux[1]+' m'))
                else:
                    Notification.objects.filter(pk = noti.pk).update(time_str=(aux[1][1]+' m'))
    return render(request, 'notifications.html', data)

def create_notification(user, action, id_property, type_p):
    users = Staff.objects.all()
    for a in users:
        if a.username_staff != user:
            alert = Notification()
            alert.user_transmitter = Staff.objects.get(username_staff=user)
            alert.user_receiver = a
            alert.type_property = type_p
            alert.id_property = id_property
            alert.action = action
            if action == 'AA':
                alert.text = 'añadió una propiedad adquirida'
                string_2 = 'view_acquisition'
                alert.link = string_2
            if action == 'AR':
                alert.text = 'añadió una propiedad rentada'
                string_2 = 'view_rent'
                alert.link = string_2
            if action == 'EA':
                alert.text = 'modificó una propiedad adquirida'
                string_2 = 'view_acquisition'
                alert.link = string_2
            if action == 'ER':
                alert.text = 'modificó una propiedad rentada'
                string_2 = 'view_rent'
                alert.link = string_2
            if action == 'SA':
                alert.text = 'habilitó el estatus de una propiedad'
                if type_p == 'AC':
                    string_2 = 'view_acquisition'
                else:
                    string_2 = 'view_rent'
                alert.link = string_2
            if action == 'SI':
                alert.text = 'deshabilitó el estatus de una propiedad'
                if type_p == 'AC':
                    string_2 = 'view_acquisition'
                else:
                    string_2 = 'view_rent'
                alert.link = string_2
            if action == 'PA':
                alert.text = 'comentó en una propiedad adquirida'
                string_2 = 'view_acquisition'
                alert.link = string_2
            if action == 'PR':
                alert.text = 'comentó en una propiedad rentada'
                string_2 = 'view_rent'
                alert.link = string_2
            if action == 'RA':
                alert.text = 'respondió en una propiedad adquirida'
                string_2 = 'view_acquisition'
                alert.link = string_2
            if action == 'RR':
                alert.text = 'respondió en una propiedad rentada'
                string_2 = 'view_rent'
                alert.link = string_2
            alert.save()